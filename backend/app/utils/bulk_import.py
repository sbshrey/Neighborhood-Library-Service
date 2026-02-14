from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..crud.books import crud_books
from ..crud.loans import crud_loans
from ..crud.users import crud_users
from ..models import Book, User
from ..schemas.books import BookCreate
from ..schemas.loans import LoanCreate
from ..schemas.users import UserCreate

SEED_DATA_DIR = Path(__file__).resolve().parents[2] / "data" / "seed_india"


@dataclass
class ImportResult:
    entity: str
    imported: int = 0
    skipped: int = 0
    errors: list[dict[str, Any]] | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "entity": self.entity,
            "imported": self.imported,
            "skipped": self.skipped,
            "errors": self.errors or [],
        }


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        normalized = value.strip()
        return normalized if normalized != "" else None
    return value


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    return int(raw)


def _parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    raw = str(value).strip()
    if not raw:
        return None
    dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    decoded = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    rows: list[dict[str, Any]] = []
    for row in reader:
        parsed = {_normalize_header(key): _normalize_cell(value) for key, value in row.items()}
        if any(value is not None for value in parsed.values()):
            rows.append(parsed)
    return rows


def _parse_xlsx(content: bytes, sheet_name: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers_row = next(iterator, None)
    if not headers_row:
        return []
    headers = [_normalize_header(value) for value in headers_row]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = {headers[index]: _normalize_cell(value) for index, value in enumerate(values)}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


async def parse_upload_rows(upload_file: UploadFile, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    filename = (upload_file.filename or "").lower()
    content = await upload_file.read()
    if not content:
        return []
    if filename.endswith(".csv"):
        return _parse_csv(content)
    if filename.endswith(".xlsx"):
        return _parse_xlsx(content, sheet_name)
    raise ValueError("Only .csv and .xlsx files are supported")


def parse_seed_csv(filename: str) -> list[dict[str, Any]]:
    path = SEED_DATA_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"Seed file not found: {path}")
    return _parse_csv(path.read_bytes())


async def import_books_rows(db: AsyncSession, rows: list[dict[str, Any]]) -> ImportResult:
    result = ImportResult(entity="books", errors=[])
    for index, row in enumerate(rows, start=2):
        try:
            isbn = row.get("isbn")
            if isbn:
                existing = (
                    await db.execute(select(Book).where(Book.isbn == str(isbn).strip()))
                ).scalar_one_or_none()
                if existing:
                    result.skipped += 1
                    continue

            payload = BookCreate(
                title=str(row.get("title") or "").strip(),
                author=str(row.get("author") or "").strip(),
                subject=row.get("subject"),
                rack_number=row.get("rack_number"),
                isbn=row.get("isbn"),
                published_year=_parse_int(row.get("published_year")),
                copies_total=_parse_int(row.get("copies_total")) or 0,
            )
            await crud_books.create(db, obj_in=payload)
            result.imported += 1
        except Exception as exc:
            result.errors.append({"row": index, "error": str(exc)})
    return result


async def import_users_rows(db: AsyncSession, rows: list[dict[str, Any]]) -> ImportResult:
    result = ImportResult(entity="users", errors=[])
    for index, row in enumerate(rows, start=2):
        try:
            email = row.get("email")
            if email:
                existing = (
                    await db.execute(select(User).where(User.email == str(email).strip()))
                ).scalar_one_or_none()
                if existing:
                    result.skipped += 1
                    continue

            payload = UserCreate(
                name=str(row.get("name") or "").strip(),
                email=str(email).strip() if email else None,
                phone=str(row.get("phone")).strip() if row.get("phone") else None,
                role=str(row.get("role") or "member").strip().lower(),
                password=str(row.get("password")).strip() if row.get("password") else None,
            )
            await crud_users.create(db, obj_in=payload)
            result.imported += 1
        except Exception as exc:
            result.errors.append({"row": index, "error": str(exc)})
    return result


async def _find_book(db: AsyncSession, row: dict[str, Any]) -> Book | None:
    book_id = _parse_int(row.get("book_id"))
    if book_id:
        return await db.get(Book, book_id)
    book_isbn = row.get("book_isbn")
    if book_isbn:
        return (
            await db.execute(select(Book).where(Book.isbn == str(book_isbn).strip()))
        ).scalar_one_or_none()
    return None


async def _find_user(db: AsyncSession, row: dict[str, Any]) -> User | None:
    user_id = _parse_int(row.get("user_id"))
    if user_id:
        return await db.get(User, user_id)
    user_email = row.get("user_email")
    if user_email:
        return (
            await db.execute(select(User).where(User.email == str(user_email).strip()))
        ).scalar_one_or_none()
    return None


async def import_loans_rows(db: AsyncSession, rows: list[dict[str, Any]]) -> ImportResult:
    result = ImportResult(entity="loans", errors=[])
    for index, row in enumerate(rows, start=2):
        try:
            book = await _find_book(db, row)
            user = await _find_user(db, row)
            if not book:
                raise ValueError("Book not found for row")
            if not user:
                raise ValueError("User not found for row")

            borrowed_at = _parse_datetime(row.get("borrowed_at")) or datetime.now(timezone.utc)
            due_at = _parse_datetime(row.get("due_at"))
            days = _parse_int(row.get("days"))

            if due_at and due_at < borrowed_at:
                raise ValueError("due_at cannot be before borrowed_at")

            if days is None:
                if due_at is not None:
                    days = max((due_at.date() - borrowed_at.date()).days, 1)
                else:
                    days = 14
            if days > settings.circulation_max_loan_days:
                raise ValueError(
                    f"Loan days cannot exceed {settings.circulation_max_loan_days} days"
                )
            if due_at is None:
                due_at = borrowed_at + timedelta(days=days)
            returned_at = _parse_datetime(row.get("returned_at"))
            if returned_at and returned_at < borrowed_at:
                raise ValueError("returned_at cannot be before borrowed_at")

            loan = await crud_loans.borrow(
                db,
                LoanCreate(book_id=book.id, user_id=user.id, days=days),
            )
            loan.borrowed_at = borrowed_at
            loan.due_at = due_at

            if returned_at:
                await crud_loans.return_loan(db, loan.id)
                loan.returned_at = returned_at

            await db.flush()
            result.imported += 1
        except Exception as exc:
            result.errors.append({"row": index, "error": str(exc)})
    return result


async def import_seed_india(db: AsyncSession) -> dict[str, Any]:
    books = await import_books_rows(db, parse_seed_csv("books.csv"))
    users = await import_users_rows(db, parse_seed_csv("users.csv"))
    loans = await import_loans_rows(db, parse_seed_csv("loans.csv"))
    return {
        "status": "completed",
        "results": [books.as_dict(), users.as_dict(), loans.as_dict()],
    }
